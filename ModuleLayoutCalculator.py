import glob
import os
import string
import os.path
import sys
import shutil
#import xlrd
import csv
import argparse
import xlsxwriter
import pyodbc
#import antigravity
import pandas as pd
import numpy as np
import matplotlib as plt
import seaborn as sns
import tkinter as tk
from tkinter.filedialog import *
from tkinter.messagebox import *
from pptx import Presentation
from pptx.util import Inches
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

# Define program related functions here.


# Precondition: function takes the length and width of the roof inputted by the user.
# Postcondition: function returns a list containing the areawise packing factor for each type of module.

def sort_fxn(given_list):

	first_index_list = []
	first_index_dict = {}
	final_list = []
	ranking = []
	first_ranking = []

	listlength = int(len(given_list)/2)

	i=0
	x=int(len(given_list)/2)
	while i < listlength:
		if given_list[i][1] > given_list[x][1]:
			first_index_dict[i] = given_list[i][1]
		else:
			first_index_dict[x] = given_list[x][1]
		x+=1
		i+=1


	first_index_dict = sorted(first_index_dict.items(), key = lambda kv:(kv[1], kv[0]), reverse = True)


	for k,v in first_index_dict:
		first_index_list.append(k)


	x=0
	while x < len(first_index_list):
		if first_index_list[x] >= listlength:
			first_ranking.append(first_index_list[x]-listlength)
			ranking.append(first_index_list[x])
		else:
			first_ranking.append(first_index_list[x])
			ranking.append(first_index_list[x]+listlength)
		x+=1

	x=0
	while x<listlength:
		final_list.append(given_list[first_ranking[x]])
		x+=1

	x=0
	while x<listlength:
		final_list.append(given_list[ranking[x]])
		x+=1

#	print(ranking)
#	print(first_ranking)
	return final_list


def new_modules(modulesfit, newpackingfactor):

	print(newpackingfactor)
	new_list = []
	order = []
	first = []
	second = []

	x = 0
	i = 0
#	print(newpackingfactor)
	counter = len(modulesfit)
	countertwo = len(modulesfit)/2
	# print(counter)
#	print(modulesfit)
	# print(newpackingfactor)

	while x < counter:
		if newpackingfactor[x][0] == modulesfit[i][0]:
			if x < countertwo:
				first.append(i)
				i=0
				x+=1
			else:
				second.append(i+countertwo)
				i=0
				x+=1
		else:
			i+=1

	print(order)
	print(len(order))
	for match in first:
		new_list.append(modulesfit[int(match)])
	for match in second:
		new_list.append(modulesfit[int(match)])
	print(second)
	print(new_list)

	return new_list


def areapackingcalculator(length, width):

	unit = dropdown()
	areapackingfactor = []
	areapackingfactor2 = []
	lengthmodulesfit = []
	widthmodulesfit = []
	lengthmodulesfit2 = []
	widthmodulesfit2 = []
	modulestoobig = []
	errorfiles = []

	modulelist = []
	lengthlistfloat = []
	widthlistfloat = []

	modfile = askopenfilename(filetypes = [("Excel Files", "*.xlsx")], title = 'Select file containing module size data.')
	print(modfile)

	conn = pyodbc.connect(r'DRIVER={Microsoft Excel Driver (*.xls, *.xlsx, *.xlsm, *.xlsb)};'+ r'DBQ='+modfile+';' r'ReadOnly=0', autocommit = True)
	conn.setdecoding(pyodbc.SQL_CHAR, encoding='utf8')
	conn.setdecoding(pyodbc.SQL_WCHAR, encoding='utf8')
	conn.setencoding(encoding='utf8')
	cursor = conn.cursor()

	for worksheet in cursor.tables():
		tablename = worksheet[3]


	cursor.execute("SELECT * FROM [{}]".format("Lookups$"))

	for row in cursor:
		modulelist.append(row[0] +' ')
		lengthlistfloat.append(row[1])
		widthlistfloat.append(row[2])

	widthlist = [str(i) for i in widthlistfloat]
	lengthlist = [str(i) for i in lengthlistfloat]
	

	arealist = np.multiply(lengthlistfloat, widthlistfloat)
	# arealist = lengthlist*widthlist
	roofarea = float(length)*float(width)

	for lengthvalue in lengthlist:
		if float(lengthvalue) <= float(length):
			lengthnummodulesfit = int(float(length)/float(lengthvalue))
			lengthmodulesfit.append(lengthnummodulesfit)
		else:
			lengthnummodulesfit = int(float(length)/float(lengthvalue))
			lengthmodulesfit.append(lengthnummodulesfit)


	for widthvalue in widthlist:
		if float(widthvalue) <= float(width):
			widthnummodulesfit = int(float(width)/float(widthvalue))
			widthmodulesfit.append(widthnummodulesfit)
		else:
			widthnummodulesfit = int(float(width)/float(widthvalue))
			widthmodulesfit.append(widthnummodulesfit)


	for lengthvalue in widthlist:
		if float(lengthvalue) <= float(length):
			lengthnummodulesfit2 = int(float(length)/float(lengthvalue))
			lengthmodulesfit2.append(lengthnummodulesfit2)
		else:
			lengthnummodulesfit2 = int(float(length)/float(lengthvalue))
			lengthmodulesfit2.append(lengthnummodulesfit2)


	for widthvalue in lengthlist:
		if float(widthvalue) <= float(width):
			widthnummodulesfit2 = int(float(width)/float(widthvalue))
			widthmodulesfit2.append(widthnummodulesfit2)
		else:
			widthnummodulesfit2 = int(float(width)/float(widthvalue))
			widthmodulesfit2.append(widthnummodulesfit2)

	areamodulesfit = np.multiply(lengthmodulesfit, widthmodulesfit)
	totalmodulearealist = np.multiply(np.array(arealist, dtype = np.int64), np.array(areamodulesfit), dtype = np.int64)

	areamodulesfit2 = np.multiply(lengthmodulesfit2, widthmodulesfit2)
	totalmodulearealist2 = np.multiply(np.array(arealist, dtype = np.int64), np.array(areamodulesfit2), dtype = np.int64)

	# print(lengthmodulesfit)
	# print(widthmodulesfit)
	# print(arealist)
	# print(areamodulesfit)
	# print(totalmodulearealist)
	# print(roofarea)

	for areavalue in totalmodulearealist:
		if float(areavalue) <= roofarea:
			areapackfactor = areavalue/roofarea
			areapackingfactor.append(str(areapackfactor))
		elif float(areavalue) >= roofarea:
			modulestoobig.append(areavalue)
		else:
			errorfiles.append(areavalue)

	areapackingfactorwithmodules = [i + j for i, j in zip(modulelist, areapackingfactor)]

	splitareapackingfactor = []

	for item in areapackingfactorwithmodules:
		splitresult = item.split()
		split_length = len(splitresult)
#		print(split_length)
		if split_length > 2:
			itemlist = [splitresult[0], splitresult[split_length-1]]
		else:
			itemlist = [splitresult[0], splitresult[1]]
		splitareapackingfactor.append(itemlist)


#here

	for areavalue in totalmodulearealist2:
		if float(areavalue) <= roofarea:
			areapackfactor = areavalue/roofarea
			areapackingfactor2.append(str(areapackfactor))
		elif float(areavalue) >= roofarea:
			modulestoobig.append(areavalue)
		else:
			errorfiles.append(areavalue)

	areapackingfactorwithmodules2 = [i + j for i, j in zip(modulelist, areapackingfactor2)]


	for item in areapackingfactorwithmodules2:
		splitresult = item.split()
		split_length = len(splitresult)
#		print(split_length)
		if split_length > 2:
			itemlist = [splitresult[0], splitresult[split_length-1]]
		else:
			itemlist = [splitresult[0], splitresult[1]]
		splitareapackingfactor.append(itemlist)

	if len(modulestoobig) != 0:
		print("These modules are too big for the given roof: ", modulestoobig)

	if len(errorfiles) != 0:
		print("An unknown lengthwise error occured: ", errorfiles)


	return splitareapackingfactor


# Precondition: function takes the length and width of the roof inputted by the user.
# Postcondition: function returns a list containing the number of modules that can fit on the given roof.


def areamodulefitcalculator(length, width):

	unit = dropdown()
	lengthmodulesfit = []
	widthmodulesfit = []
	lengthmodulesfit2 = []
	widthmodulesfit2 = []
	modulestoobig = []
	errorfiles = []

	modulelist = []
	lengthlistfloat = []
	widthlistfloat = []

	modfile = askopenfilename(filetypes = [("Excel Files", "*.xlsx")], title = 'Select file containing module size data.')


	conn = pyodbc.connect(r'DRIVER={Microsoft Excel Driver (*.xls, *.xlsx, *.xlsm, *.xlsb)};'+ r'DBQ='+modfile+';' r'ReadOnly=0', autocommit = True)
	conn.setdecoding(pyodbc.SQL_CHAR, encoding='utf8')
	conn.setdecoding(pyodbc.SQL_WCHAR, encoding='utf8')
	conn.setencoding(encoding='utf8')
	cursor = conn.cursor()

	for worksheet in cursor.tables():
		tablename = worksheet[3]

	cursor.execute("SELECT * FROM [{}]".format("Lookups$"))

	for row in cursor:
		modulelist.append(row[0] +' ')
		lengthlistfloat.append(row[1])
		widthlistfloat.append(row[2])

	widthlist = [str(i) for i in widthlistfloat]
	lengthlist = [str(i) for i in lengthlistfloat]
	

	arealist = np.multiply(lengthlistfloat, widthlistfloat)

	# arealist = lengthlist*widthlist
	roofarea = float(length)*float(width)

	for lengthvalue in lengthlist:
		if float(lengthvalue) <= float(length):
			lengthnummodulesfit = int(float(length)/float(lengthvalue))
			lengthmodulesfit.append(lengthnummodulesfit)
		else:
			lengthnummodulesfit = int(float(length)/float(lengthvalue))
			lengthmodulesfit.append(lengthnummodulesfit)

	for widthvalue in widthlist:
		if float(widthvalue) <= float(width):
			widthnummodulesfit = int(float(width)/float(widthvalue))
			widthmodulesfit.append(widthnummodulesfit)
		else:
			widthnummodulesfit = int(float(width)/float(widthvalue))
			widthmodulesfit.append(widthnummodulesfit)

	for lengthvalue in widthlist:
		if float(lengthvalue) <= float(length):
			lengthnummodulesfit = int(float(length)/float(lengthvalue))
			lengthmodulesfit2.append(lengthnummodulesfit)
		else:
			lengthnummodulesfit = int(float(length)/float(lengthvalue))
			lengthmodulesfit2.append(lengthnummodulesfit)

	for widthvalue in lengthlist:
		if float(widthvalue) <= float(width):
			widthnummodulesfit = int(float(width)/float(widthvalue))
			widthmodulesfit2.append(widthnummodulesfit)
		else:
			widthnummodulesfit = int(float(width)/float(widthvalue))
			widthmodulesfit2.append(widthnummodulesfit)

	areamodulesfit = np.multiply(lengthmodulesfit, widthmodulesfit)
	stringareamodulesfit = [str(i) for i in areamodulesfit]

	areamodulesfitwithmodules = [i + j for i, j in zip(modulelist, stringareamodulesfit)]

	areamodulesfit2 = np.multiply(lengthmodulesfit2, widthmodulesfit2)
	stringareamodulesfit2 = [str(i) for i in areamodulesfit2]

	areamodulesfitwithmodules2 = [i + j for i, j in zip(modulelist, stringareamodulesfit2)]

	splitnummodulesfit = []

	for item in areamodulesfitwithmodules:
		splitresult = item.split()
		split_length = len(splitresult)
#		print(split_length)
		if split_length > 2:
			itemlist = [splitresult[0], splitresult[split_length-1]]
		else:
			itemlist = [splitresult[0], splitresult[1]]
		splitnummodulesfit.append(itemlist)

	for item in areamodulesfitwithmodules2:
		splitresult = item.split()
		split_length = len(splitresult)
#		print(split_length)
		if split_length > 2:
			itemlist = [splitresult[0], splitresult[split_length-1]]
		else:
			itemlist = [splitresult[0], splitresult[1]]
		splitnummodulesfit.append(itemlist)



	return splitnummodulesfit





def getpackingfactor():
	rooflength = entry1.get()
	roofwidth = entry2.get()
	unit = dropdown()

	unitconv = 1

	if unit == "in":
		roofwidth = float(roofwidth)*25.4
		rooflength= float(rooflength)*25.4
		unitconv = 25.4

	if unit == "cm":
		roofwidth = float(roofwidth)*10
		rooflength = float(rooflength)*10
		unitconv = 10

	if unit == "ft":
		roofwidth = float(roofwidth)*304.8
		rooflength = float(rooflength)*304.8
		unitconv = 304.8

	if unit == "m":
		roofwidth = float(roofwidth)*1000
		rooflength = float(rooflength)*1000
		unitconv = 1000

	if unit == "yd":
		roofwidth = float(roofwidth)*914.4
		rooflength = float(rooflength)*914.4
		unitconv = 914.4


	#print(rooflength)

	packingfactor = areapackingcalculator(rooflength, roofwidth)

	newpackingfactor = sort_fxn(packingfactor)
	print(newpackingfactor)

	rooflength = float(rooflength)/unitconv
	roofwidth = float(roofwidth)/unitconv

	location = askdirectory(title = "Select location to export packing factors to.")
	os.chdir(location)

	filename = entry3.get()

	with xlsxwriter.Workbook(str(filename)+".xlsx") as workbook:
		worksheet = workbook.add_worksheet('Packing Factor')

		worksheet.write(0, 0, "Module Serial Number")
		worksheet.write(0, 1, "Module Packing Factor (L. Panel x L. Roof)")
		worksheet.write(0, 2, "Module Packing Factor (L.Panel x W. Roof)")
		worksheet.write(0, 3, "Roof Dimensions")
		worksheet.write(1, 3, "Length "+unit)
		worksheet.write(2, 3, "Width "+unit)
		worksheet.write(1, 4, float(rooflength))
		worksheet.write(2, 4, float(roofwidth))

		row = 1
		col = 0
		row2 = 1
		#print(packfactor)
		lengthpack = (len(newpackingfactor)/2)+1

		for module, packfactor in newpackingfactor:
			if row <lengthpack:
				worksheet.write(row, col, module)
				worksheet.write(row, col + 1, float(packfactor))
				row += 1
			else:
				worksheet.write(row2, col + 2, float(packfactor))				
				row2 += 1


		worksheet.conditional_format('B2:B28',{'type':'3_color_scale', 'min_color': 'red', 'mid_color': '#FFCC00', 'max_color': '#008000',
			'min_value':'0', 'mid_value': '0.5', 'max_value': '1'})

		worksheet.conditional_format('C2:B28',{'type':'3_color_scale', 'min_color': 'red', 'mid_color': '#FFCC00', 'max_color': '#008000',
			'min_value':'0', 'mid_value': '0.5', 'max_value': '1'})

		#workbook.close()

	label5 = tk.Label(root, text = "The packing factor of each module is exported to an Excel spreadsheet located in the selected folder.")
	canvas1.create_window(350, 380, window = label5)

def getnummodules():
	rooflength = entry1.get()
	roofwidth = entry2.get()
	unit = dropdown()

	unitconv = 1

	if unit == "in":
		roofwidth = float(roofwidth)*25.4
		rooflength= float(rooflength)*25.4
		unitconv = 25.4

	if unit == "cm":
		roofwidth = float(roofwidth)*10
		rooflength = float(rooflength)*10
		unitconv = 10

	if unit == "ft":
		roofwidth = float(roofwidth)*304.8
		rooflength = float(rooflength)*304.8
		unitconv = 304.8

	if unit == "m":
		roofwidth = float(roofwidth)*1000
		rooflength = float(rooflength)*1000
		unitconv = 1000

	if unit == "yd":
		roofwidth = float(roofwidth)*914.4
		rooflength = float(rooflength)*914.4
		unitconv = 914.4


	modulesfit = areamodulefitcalculator(rooflength, roofwidth)

	rooflength = float(rooflength)/unitconv
	roofwidth = float(roofwidth)/unitconv


	location = askdirectory(title = "Select location to export number of modules fit to.")
	os.chdir(location)

	filename = entry3.get()

	with xlsxwriter.Workbook(str(filename)+".xlsx") as workbook:
		worksheet = workbook.add_worksheet()

		worksheet.write(0, 0, "Module Serial Number")
		worksheet.write(0, 1, "Number of Modules that Fit on Roof (L. Panel x L. Roof)")
		worksheet.write(0, 2, "Number of Modules that Fit on Roof (L. Panel x W. Roof)")	
		worksheet.write(0, 3, "Roof Dimensions")
		worksheet.write(1, 3, "Length ("+unit+")")
		worksheet.write(2, 3, "Width ("+unit+")")
		worksheet.write(1, 4, float(rooflength))
		worksheet.write(2, 4, float(roofwidth))

		row = 1
		col = 0
		row2 = 1
		maxnum = 0
		minnum = 0
		maxnum2 = 0
		minnum2 = 0

		lengthmod = (len(modulesfit)/2)+1
		for module, numfit in modulesfit:
			if row < lengthmod:
				worksheet.write(row, col, module)
				worksheet.write(row, col + 1, float(numfit))
				if float(numfit) < minnum:
					minnum = float(numfit)
				if float(numfit) > maxnum:
					maxnum = float(numfit)
				row += 1
			else:
				worksheet.write(row2, col + 2, float(numfit))
				if float(numfit) < minnum2:
					minnum2 = float(numfit)
				if float(numfit) > maxnum2:
					maxnum2 = float(numfit)
				row2 += 1

		midnum = (minnum +maxnum)/2
		midnum2 = (minnum2+maxnum2)/2

		worksheet.conditional_format('B2:B28',{'type':'3_color_scale', 'min_color': 'red', 'mid_color': '#FFCC00', 'max_color': '#008000',
			'min_value':'minnum', 'mid_value': 'midnum', 'max_value': 'maxnum'})
		worksheet.conditional_format('C2:B28',{'type':'3_color_scale', 'min_color': 'red', 'mid_color': '#FFCC00', 'max_color': '#008000',
			'min_value':'minnum2', 'mid_value': 'midnum2', 'max_value': 'maxnum2'})

		#workbook.close()

	label6 = tk.Label(root, text = "The number of modules that fit is exported to an Excel spreadheet located in the selected folder.")
	canvas1.create_window(350, 380, window = label6)

def packing_and_fit():
	rooflength = entry1.get()
	roofwidth = entry2.get()
	unit = dropdown()

	unitconv = 1

	if unit == "in":
		roofwidth = float(roofwidth)*25.4
		rooflength= float(rooflength)*25.4
		unitconv = 25.4

	if unit == "cm":
		roofwidth = float(roofwidth)*10
		rooflength = float(rooflength)*10
		unitconv = 10

	if unit == "ft":
		roofwidth = float(roofwidth)*304.8
		rooflength = float(rooflength)*304.8
		unitconv = 304.8

	if unit == "m":
		roofwidth = float(roofwidth)*1000
		rooflength = float(rooflength)*1000
		unitconv = 1000

	if unit == "yd":
		roofwidth = float(roofwidth)*914.4
		rooflength = float(rooflength)*914.4
		unitconv = 914.4



	packingfactor = areapackingcalculator(rooflength, roofwidth)

	#print(packingfactor)

	newpackingfactor = sort_fxn(packingfactor)
	#print(newpackingfactor)

	rooflength = float(rooflength)/unitconv
	roofwidth = float(roofwidth)/unitconv

	location = askdirectory(title = "Select location to export packing factors to.")
	os.chdir(location)


	filename = entry3.get()

	rooflength = entry1.get()
	roofwidth = entry2.get()
	unit = dropdown()

	unitconv = 1

	if unit == "in":
		roofwidth = float(roofwidth)*25.4
		rooflength= float(rooflength)*25.4
		unitconv = 25.4

	if unit == "cm":
		roofwidth = float(roofwidth)*10
		rooflength = float(rooflength)*10
		unitconv = 10

	if unit == "ft":
		roofwidth = float(roofwidth)*304.8
		rooflength = float(rooflength)*304.8
		unitconv = 304.8

	if unit == "m":
		roofwidth = float(roofwidth)*1000
		rooflength = float(rooflength)*1000
		unitconv = 1000

	if unit == "yd":
		roofwidth = float(roofwidth)*914.4
		rooflength = float(rooflength)*914.4
		unitconv = 914.4




	modulesfit = areamodulefitcalculator(rooflength, roofwidth)

	rooflength = float(rooflength)/unitconv
	roofwidth = float(roofwidth)/unitconv

	filename = entry3.get()

	with xlsxwriter.Workbook(str(filename)+".xlsx") as workbook:
		worksheet = workbook.add_worksheet('Packing Factor')

		worksheet.write(0, 0, "Module Serial Number")
		worksheet.write(0, 1, "Module Packing Factor (L. Panel x L. Roof)")
		worksheet.write(0, 2, "Module Packing Factor (L.Panel x W. Roof)")
		worksheet.write(0, 3, "Roof Dimensions")
		worksheet.write(1, 3, "Length "+unit)
		worksheet.write(2, 3, "Width "+unit)
		worksheet.write(1, 4, float(rooflength))
		worksheet.write(2, 4, float(roofwidth))

		row = 1
		col = 0
		row2 = 1


		lenpack = (len(newpackingfactor)/2)+1
		for module, packfactor in newpackingfactor:
			if row <lenpack:
				worksheet.write(row, col, module)
				worksheet.write(row, col + 1, float(packfactor))
				row += 1
			else:
				worksheet.write(row2, col + 2, float(packfactor))				
				row2 += 1
 

		worksheet.conditional_format('B2:B50',{'type':'3_color_scale', 'min_color': 'red', 'mid_color': '#FFCC00', 'max_color': '#008000',
			'min_value':'0', 'mid_value': '0.5', 'max_value': '1'})
		worksheet.conditional_format('C2:B50',{'type':'3_color_scale', 'min_color': 'red', 'mid_color': '#FFCC00', 'max_color': '#008000',
			'min_value':'0', 'mid_value': '0.5', 'max_value': '1'})


#		print(modulesfit)
#		print(newpackingfactor)
		newmodulesfit = new_modules(modulesfit, newpackingfactor)
		print(newmodulesfit)


		worksheet = workbook.add_worksheet('Modules Fit')

		worksheet.write(0, 0, "Module Serial Number")
		worksheet.write(0, 1, "Number of Modules that Fit on Roof (L. Panel x L. Roof)")
		worksheet.write(0, 2, "Number of Modules that Fit on Roof (L. Panel x W. Roof)")		
		worksheet.write(0, 3, "Roof Dimensions")
		worksheet.write(1, 3, "Length ("+unit+")")
		worksheet.write(2, 3, "Width ("+unit+")")
		worksheet.write(1, 4, float(rooflength))
		worksheet.write(2, 4, float(roofwidth))

		row = 1
		col = 0
		row2 = 1


		#print(modulesfit)
		lengthmod = (len(modulesfit)/2)+1
		for module, numfit in newmodulesfit:
			if row < lengthmod:
				worksheet.write(row, col, module)
				worksheet.write(row, col + 1, float(numfit))
				row += 1
			else:
				worksheet.write(row2, col + 2, float(numfit))
				row2 += 1

		#workbook.close()

	label6 = tk.Label(root, text = "The requested data has been exported to an Excel spreadheet located in the selected folder.")
	canvas1.create_window(350, 380, window = label6)




# Part 1: Prompt user to input the length and width of the rectangular roof in meters.


root = tk.Tk()

canvas1 = tk.Canvas(root, width = 700, height = 500)
canvas1.pack()

entry1 = tk.Entry(root)
canvas1.create_window(350, 100, window = entry1)

label1 = tk.Label(text = " Roof Length")
canvas1.create_window(220, 100, window = label1)

entry2 = tk.Entry(root)
canvas1.create_window(350, 140, window = entry2)

label2 = tk.Label(text = "Roof Width")
canvas1.create_window(220, 140, window = label2)


tkvar = tk.StringVar(root)
choices = {'mm', 'cm', 'in', 'ft', 'yd', 'm'}
tkvar.set('mm')

popUpMenu = tk.OptionMenu(root, tkvar, *choices)
unitlabel = tk.Label(text = "Units (Ex- cm; in; ft;)")
canvas1.create_window(240,180,window =unitlabel)
canvas1.create_window(350,180, window = popUpMenu)

def dropdown(*args):
	return tkvar.get()

tkvar.trace('w', dropdown)


entry3 = tk.Entry(root)
canvas1.create_window(450, 220, window = entry3)

label3 = tk.Label(text = "Packing Factor Spreadsheet Filename")
canvas1.create_window(240, 220, window = label3)

#entry4 = tk.Entry(root)
#canvas1.create_window(450, 260, window = entry4)

#label4 = tk.Label(text = "Number of Modules Fit Spreadsheet Filename")
#canvas1.create_window(240, 260, window = label4)


# Part 2: Return a list of packing factors for all modules for both orientations (no mixing of anything).


button1 = tk.Button(text = "Packing Factor of All Modules", command = getpackingfactor)
canvas1.create_window(350, 260, window = button1)

button2 = tk.Button(text = "Number of Modules that Fit", command = getnummodules)
canvas1.create_window(350, 300, window = button2)

button3 = tk.Button(text = "Packing Factor and Modules that Fit", command = packing_and_fit)
canvas1.create_window(350,340, window = button3)

root.mainloop()

